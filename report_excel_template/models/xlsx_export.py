# -*- coding: utf-8 -*-
import datetime
import os
import logging
import base64
from io import BytesIO
import time
from datetime import date, datetime as dt
from odoo.tools.float_utils import float_compare
from odoo import models, fields, api, _
from odoo.tools.safe_eval import safe_eval
from odoo.exceptions import ValidationError
from . import common as co
import re
from copy import copy

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug('Cannot import "openpyxl". Please make sure it is installed.')


class XLSXExport(models.AbstractModel):
    _name = 'xlsx.export'
    _description = 'Excel Export AbstractModel'

    @api.model
    def get_eval_context(self, model, record, value):
        eval_context = {'float_compare': float_compare,
                        #'time': time,
                        'datetime': dt,
                        'date': date,
                        'value': value,
                        'object': record,
                        'model': self.env[model],
                        'env': self.env,
                        'context': self._context,
                        }
        return eval_context

    @api.model
    def _get_line_vals(self, record, line_field, fields):
        """ Get values of this field from record set and return as dict of vals
            - record: main object
            - line_field: rows object, i.e., line_ids
            - fields: fields in line_ids, i.e., partner_id.display_name
        """
        line_field, max_row = co.get_line_max(line_field)
        line_field = line_field.replace('_CONT_', '')  # Remove _CONT_ if any
        lines = record[line_field]
        if max_row > 0 and len(lines) > max_row:
            raise Exception(
                _('Records in %s exceed max records allowed') % line_field)
        vals = dict([(field, []) for field in fields])  # value and do_style
        # Get field condition & aggre function
        field_cond_dict = {}
        aggre_func_dict = {}
        field_style_dict = {}
        style_cond_dict = {}
        pair_fields = []  # I.e., ('debit${value and . or .}@{sum}', 'debit')
        for field in fields:
            temp_field, eval_cond = co.get_field_condition(field)
            eval_cond = eval_cond or 'value or ""'
            temp_field, field_style = co.get_field_style(temp_field)
            temp_field, style_cond = co.get_field_style_cond(temp_field)
            raw_field, aggre_func = co.get_field_aggregation(temp_field)
            # Dict of all special conditions
            field_cond_dict.update({field: eval_cond})
            aggre_func_dict.update({field: aggre_func})
            field_style_dict.update({field: field_style})
            style_cond_dict.update({field: style_cond})
            # --
            pair_fields.append((field, raw_field))
        for line in lines:
            for field in pair_fields:  # (field, raw_field)
                value = self._get_field_data(field[1], line)
                eval_cond = field_cond_dict[field[0]]
                eval_context = \
                    self.get_eval_context(line._name, line, value)
                if eval_cond:
                    value = safe_eval(eval_cond, eval_context)
                # style w/Cond takes priority
                style_cond = style_cond_dict[field[0]]
                style = self._eval_style_cond(line._name, line,
                                              value, style_cond)
                if style is None:
                    style = False  # No style
                elif style is False:
                    style = field_style_dict[field[0]]  # Use default style
                vals[field[0]].append((value, style))
        return (vals, aggre_func_dict,)

    @api.model
    def _eval_style_cond(self, model, record, value, style_cond):
        eval_context = self.get_eval_context(model, record, value)
        field = style_cond = style_cond or '#??'
        styles = {}
        for i in range(style_cond.count('#{')):
            i += 1
            field, style = co.get_field_style(field)
            styles.update({i: style})
            style_cond = style_cond.replace('#{%s}' % style, str(i))
        if not styles:
            return False
        res = safe_eval(style_cond, eval_context)
        if res is None or res is False:
            return res
        return styles[res]

    @api.model
    def _fill_workbook_data(self, workbook, record, data_dict):
        """ Fill data from record with style in data_dict to workbook """
        if not record or not data_dict:
            return
        try:
            for sheet_name in data_dict:
                ws = data_dict[sheet_name]
                st = False
                if isinstance(sheet_name, str):
                    st = co.openpyxl_get_sheet_by_name(workbook, sheet_name)
                elif isinstance(sheet_name, int):
                    if sheet_name > len(workbook.worksheets):
                        raise Exception(_('Not enough worksheets'))
                    st = workbook.worksheets[sheet_name - 1]
                if not st:
                    raise ValidationError(
                        _('Sheet %s not found') % sheet_name)
                # Fill data, header and rows
                self._fill_head(ws, st, record)
                self._fill_lines(ws, st, record)
        except KeyError as e:
            raise ValidationError(_('Key Error\n%s') % e)
        except IllegalCharacterError as e:
            raise ValidationError(
                _('IllegalCharacterError\n'
                  'Some exporting data contain special character\n%s') % e)
        except Exception as e:
            raise ValidationError(
                _('Error filling data into Excel sheets\n%s') % e)

    @api.model
    def _get_field_data(self, _field, _line):
        """ Get field data, and convert data type if needed """
        if not _field:
            return None
        line_copy = _line
        for f in _field.split('.'):
            line_copy = line_copy[f]
        if isinstance(line_copy, str):
            line_copy = line_copy.encode('utf-8')
        return line_copy

    @api.model
    def _fill_head(self, ws, st, record):
        for rc, field in ws.get('_HEAD_', {}).items():
            tmp_field, eval_cond = co.get_field_condition(field)
            eval_cond = eval_cond or 'value or ""'
            tmp_field, field_style = co.get_field_style(tmp_field)
            tmp_field, style_cond = co.get_field_style_cond(tmp_field)
            value = tmp_field and self._get_field_data(tmp_field, record)
            # Eval
            eval_context = self.get_eval_context(record._name,
                                                 record, value)
            if eval_cond:
                value = safe_eval(eval_cond, eval_context)
            if value is not None:
                st[rc] = value
            fc = not style_cond and True or \
                safe_eval(style_cond, eval_context)
            if field_style and fc:  # has style and pass style_cond
                styles = self.env['xlsx.styles'].get_openpyxl_styles()
                co.fill_cell_style(st[rc], field_style, styles)

    @api.model
    def _fill_lines(self, ws, st, record):
        line_fields = list(ws)
        if '_HEAD_' in line_fields:
            line_fields.remove('_HEAD_')
        cont_row = 0  # last data row to continue
        for line_field in line_fields:
            fields = ws.get(line_field, {}).values()
            vals, func = self._get_line_vals(record, line_field, fields)
            is_cont = '_CONT_' in line_field and True or False  # continue row
            cont_set = 0
            rows_inserted = False  # flag to insert row
            for rc, field in ws.get(line_field, {}).items():
                col, row = co.split_row_col(rc)  # starting point
                # Case continue, start from the last data row
                if is_cont and not cont_set:  # only once per line_field
                    cont_set = cont_row + 1
                if is_cont:
                    row = cont_set
                    rc = '%s%s' % (col, cont_set)
                i = 0
                new_row = 0
                new_rc = False
                row_count = len(vals[field])
                # Insert rows to preserve total line
                if not rows_inserted:
                    rows_inserted = True
                    st.insert_rows(row+1, amount=row_count-1)
                # --
                for (row_val, style) in vals[field]:
                    new_row = row + i
                    new_rc = '%s%s' % (col, new_row)
                    row_val = co.adjust_cell_formula(row_val, i)
                    if row_val not in ('None', None):
                        st[new_rc] = co.str_to_number(row_val)
                    if style:
                        styles = self.env['xlsx.styles'].get_openpyxl_styles()
                        co.fill_cell_style(st[new_rc], style, styles)
                    i += 1
                # Add footer line if at least one field have sum
                f = func.get(field, False)
                if f and new_row > 0:
                    new_row += 1
                    f_rc = '%s%s' % (col, new_row)
                    st[f_rc] = '=%s(%s:%s)' % (f, rc, new_rc)
                    co.fill_cell_style(st[f_rc], style, styles)
                cont_row = cont_row < new_row and new_row or cont_row
        return

    @api.model
    def export_xlsx(self, template, res_model, res_id):
        if template.model_id.model != res_model:
            raise ValidationError(_("Template's model mismatch"))
        data_dict = co.literal_eval(template.instruction.strip())
        export_dict = data_dict.get('__EXPORT__', False)
        out_name = template.name
        if not export_dict:  # If there is not __EXPORT__ formula, just export
            out_name = template.name
            out_file = template.file_excel_template
            return (out_file, out_name)
        # Prepare temp file (from now, only xlsx file works for openpyxl)
        decoded_data = base64.decodestring(template.file_excel_template)
        ConfParam = self.env['ir.config_parameter'].sudo()
        ptemp = ConfParam.get_param('path_temp_file') or '/tmp'
        stamp = dt.utcnow().strftime('%H%M%S%f')[:-3]
        ftemp = '%s/temp%s.xlsx' % (ptemp, stamp)
        f = open(ftemp, 'wb')
        f.write(decoded_data)
        f.seek(0)
        f.close()
        # Workbook created, temp file removed
        wb = load_workbook(ftemp)
        os.remove(ftemp)
        # Start working with workbook
        record = res_model and self.env[res_model].browse(res_id) or False
        self._fill_workbook_data(wb, record, export_dict)
        # Return file as .xlsx
        content = BytesIO()
        wb.save(content)
        content.seek(0)  # Set index to 0, and start reading
        out_file = content.read()
        content.close()
        return out_file

    def _fill_workbook_data_report(self, workbook, datas, inputs, report):
        if report.id == False:
            return
        try:
            st = False
            sheet_name = report.sheet_name
            if isinstance(sheet_name, str):
                st = co.openpyxl_get_sheet_by_name(workbook, sheet_name)
            elif isinstance(sheet_name, int):
                if sheet_name > len(workbook.worksheets):
                    raise Exception(_('Not enough worksheets'))
                st = workbook.worksheets[sheet_name - 1]
            if not st:
                raise ValidationError(
                    _('Sheet %s not found') % sheet_name)
            #Header
            for input in report.input_ids:
                if input.excel_cell == '':
                    continue
                if input.excel_cell == False:
                    continue
                tmp_field, eval_cond = co.get_field_condition(input.field_cond)
                eval_cond = eval_cond or 'value or ""'
                tmp_field, field_style = co.get_field_style(input.style)
                tmp_field, style_cond = co.get_field_style_cond(input.style_cond)
                tmp_field = input.name
                value = inputs[tmp_field].encode('utf-8')
                if input.type == 'many2one':
                    input_object = self.env[input.model_id.model].search([('id','=', int(value))], limit=1)
                    eval_context = self.get_eval_context(input_object._name, input_object, value)
                elif input.type == 'date':
                    value = datetime.datetime.strptime(inputs[tmp_field], '%Y-%m-%d')
                    lang_id = self.env['res.lang'].search([('code','=', self.env.lang)], limit=1)
                    value = value.strftime(lang_id.date_format)
                    eval_context = self.get_eval_context(report._name, report, value)
                else:
                    eval_context = self.get_eval_context(report._name,report, value)
                if eval_cond:
                    value = safe_eval(eval_cond, eval_context)
                if value is not None:
                    st[input.excel_cell] = value
                fc = not style_cond and True or \
                     safe_eval(style_cond, eval_context)
                if field_style and fc:  # has style and pass style_cond
                    styles = self.env['xlsx.styles'].get_openpyxl_styles()
                    co.fill_cell_style(st[input.excel_cell], field_style, styles)
            #lines
            i = 0
            total_row = len(datas)
            cell_template = False
            for data in datas:
                insert_row = True
                if report.type == 'sql':
                    labels = report.label_ids
                else:
                    labels = report.field_ids
                for label in labels:
                    if label.excel_cell == False:
                        continue
                    tmp_f, eval_cond = co.get_field_condition(label.field_cond)
                    eval_cond = eval_cond or 'value or ""'
                    tmp_f, field_style = co.get_field_style(label.style)
                    tmp_f, style_cond = co.get_field_style_cond(label.style_cond)
                    col, row = co.split_row_col(label.excel_cell)
                    row = int(row) + i
                    if total_row - 1 > i:
                        if insert_row:
                            insert_row = False
                            st.insert_rows(row)
                    if label.numerical_order:
                        value = i + 1
                    else:
                        if report.type == 'sql':
                            value = data[label.name]
                        else:
                            value = data[label.label]
                            if isinstance(value, datetime.datetime):
                                lang_id = self.env['res.lang'].search([('code','=', self.env.user.lang)],limit=1)
                                value = value.strftime(lang_id.date_format + ' ' + lang_id.time_format)
                            if isinstance(value, datetime.date):
                                lang_id = self.env['res.lang'].search([('code', '=', self.env.user.lang)], limit=1)
                                value = value.strftime(lang_id.date_format)
                            if isinstance(value, tuple):
                                value = value[1]

                    cell = col + str(row)
                    eval_context = self.get_eval_context(report._name, report, value)
                    if eval_cond:
                        value = safe_eval(eval_cond, eval_context)
                    if value is not None:
                        st[cell] = value
                    fc = not style_cond and True or \
                         safe_eval(style_cond, eval_context)
                    if field_style and fc:  # has style and pass style_cond
                        styles = self.env['xlsx.styles'].get_openpyxl_styles()
                        co.fill_cell_style(st[cell], field_style, styles)
                    if label.sum and total_row-1 == i:
                        cell_sum = '= SUM(' + label.excel_cell + ':' + cell + ')'
                        st[col + str(row+1)] = cell_sum
                        if field_style and fc:
                            styles = self.env['xlsx.styles'].get_openpyxl_styles()
                            co.fill_cell_style(st[col + str(row+1)], field_style, styles)
                i = i + 1

        except KeyError as e:
            raise ValidationError(_('Key Error\n%s') % e)
        except IllegalCharacterError as e:
            raise ValidationError(
                _('IllegalCharacterError\n'
                  'Some exporting data contain special character\n%s') % e)
        except Exception as e:
            raise ValidationError(
                _('Error filling data into Excel sheets\n%s') % e)

    @api.model
    def export_report_excel(self, report_id, datas, inputs):
        # Prepare temp file (from now, only xlsx file works for openpyxl)
        decoded_data = base64.decodestring(report_id.file_excel_template)
        ConfParam = self.env['ir.config_parameter'].sudo()
        ptemp = ConfParam.get_param('path_temp_file') or '/tmp'
        stamp = dt.utcnow().strftime('%H%M%S%f')[:-3]
        ftemp = '%s/temp%s.xlsx' % (ptemp, stamp)
        f = open(ftemp, 'wb')
        f.write(decoded_data)
        f.seek(0)
        f.close()
        # Workbook created, temp file removed
        wb = load_workbook(ftemp)
        os.remove(ftemp)
        # Start working with workbook
        self._fill_workbook_data_report(wb, datas, inputs, report_id)
        content = BytesIO()
        wb.save(content)
        content.seek(0)
        return content,report_id.name